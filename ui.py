import tkinter as tk
from tkinter import ttk
import time
import threading
from tkinter import messagebox
import new_frcnn_test
import os
import queue
import new_frcnn_test
import xml.etree.ElementTree as ET  # Import XML module
import status_shared
import openpyxl

# Function to update the Excel file with the works order number and relevant counts
def update_excel(works_order_no, fp_count, fn_count):
    file_path = "WorksOrder_issues.xlsx"
    try:
        if not os.path.exists(file_path):
            # Create a new workbook and add a worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["wk_no", "FP", "FN"])  # Add headers
        else:
            # Load existing workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

        # Find the row corresponding to the works order number
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == works_order_no:
                # Update counts
                ws.cell(row=row[0].row, column=2).value = (ws.cell(row=row[0].row, column=2).value or 0) + fp_count
                ws.cell(row=row[0].row, column=3).value = (ws.cell(row=row[0].row, column=3).value or 0) + fn_count
                break
        else:
            # Add a new row for the works order number
            ws.append([works_order_no, fp_count, fn_count])

        wb.save(file_path)

        print("Updated Excel file with works order number:", works_order_no)
        print("FP Count:", fp_count)
        print("FN Count:", fn_count)

    except Exception as e:
        print("An error occurred while updating the Excel file:", e)

# Function to monitor the output file and update the UI
def monitor_output_file(file_path, queue):
    processing_image = False
    with open(file_path, "r") as file:
        file.seek(0, os.SEEK_END)  # Start at the end of the file
        while True:
            line = file.readline()
            if not line:
                time.sleep(0.1)
                continue
            if "====" in line:  # Delimiter for image processing blocks
                processing_image = not processing_image  # Toggle processing state
                if not processing_image:
                    queue.put("grey")  # Reset to grey when block ends
                continue
            if processing_image:
                if "Missing" in line:
                    queue.put("red")
                elif "Everything looks good!" in line:
                    queue.put("green")

# Function to update the UI based on status
def update_ui_based_on_status():
    status_value = status_shared.status['value']
    if status_value == 1:
        status_indicator.config(bg="red")
    elif status_value == 2:
        status_indicator.config(bg="green")
    else:
        status_indicator.config(bg="grey")
    update_accuracy_label
    root.after(100, update_ui_based_on_status)  # Schedule this function to run every 100ms

# Function to update the UI (status indicator color)
def update_callback(color):
    print("Updating status indicator with color:", color)
    status_indicator.config(bg=color)

def on_take_photo_button_click():
    print("Button clicked, now taking photo...")
    try:
        # Retrieve threshold values from entry widgets
        label_thresholds = [label_threshold1.get(), label_threshold2.get(), label_threshold3.get()]
        screw_thresholds = [screw_threshold1.get(), screw_threshold2.get(), screw_threshold3.get()]

        # Convert thresholds to integers
        try:
            label_thresholds = [int(label) for label in label_thresholds]
            screw_thresholds = [int(screw) for screw in screw_thresholds]
        except ValueError:
            messagebox.showerror("Error", "Thresholds must be integers")
            return

        print("Label Thresholds:", label_thresholds)  # Add this line to print label thresholds
        print("Screw Thresholds:", screw_thresholds)  # Add this line to print screw thresholds

        global total_images
        # Call take_photo with threshold values
        new_frcnn_test.take_photo(*label_thresholds, *screw_thresholds, total_images)
        # Increment the total images count after each set of images is processed
        #global total_images
        #total_images += 3

    except Exception as e:
        print(f"Error taking photo: {e}")


def clear_entry(event, entry, default_text):
    if entry.get() == default_text:
        entry.delete(0, tk.END)
        entry.config(foreground="black")  # Change text color to black

def restore_default_text(event, entry, default_text):
    if not entry.get():
        entry.insert(0, default_text)
        entry.config(foreground="black")  # Change text color to black


# Initialize variables
total_images = 0
mistake_count = 0
false_positive_count = 0
false_negative_count = 0

# Function to report a mistake and update the accuracy label
def report_mistake():
    global mistake_count
    mistake_count += 1
    update_accuracy_label()

#def update_accuracy_label():
    #global total_images, mistake_count
    #accuracy = 100 * (1 - mistake_count / total_images)
    #accuracy_label.config(text=f"Accuracy: {accuracy:.2f}%")

def update_accuracy_label():
    global total_images, mistake_count
    total_images = status_shared.total_images
    if total_images == 0:
        accuracy = 0  # Set accuracy to 0 if no images have been processed yet
    else:
        accuracy = 100 * (1 - mistake_count / total_images)
    accuracy_label.config(text=f"Accuracy: {accuracy:.2f}%")


# Function to report a false positive
def report_false_positive():
    global false_positive_count
    works_order_no = works_input.get()  # Fetch works order number from entry field
    if works_order_no.strip() == "":
        messagebox.showerror("Error", "Works Order No. cannot be empty")
        return
    false_positive_count += 1
    update_fp_label()
    report_mistake()
    update_excel(works_order_no, 1, 0)  # Update xlsx with works order number and increment FP count

# Function to report a false negative
def report_false_negative():
    global false_negative_count
    works_order_no = works_input.get()  # Fetch works order number from entry field
    if works_order_no.strip() == "":
        messagebox.showerror("Error", "Works Order No. cannot be empty")
        return
    false_negative_count += 1
    update_fn_label()
    report_mistake()
    update_excel(works_order_no, 0, 1)  # Update xlsx with works order number and increment FN count
    
# Function to update the false positive label
def update_fp_label():
    fp_label.config(text=f"FP: {false_positive_count}")

# Function to update the false negative label
def update_fn_label():
    fn_label.config(text=f"FN: {false_negative_count}")

def main():
    print("Starting the GUI application...")
    global root, label_threshold1, label_threshold2, label_threshold3, screw_threshold1, screw_threshold2, screw_threshold3, error_label, status_indicator
    root = tk.Tk()
    root.title("Photo Capture")
    update_queue = queue.Queue()

    # Global variable for accuracy label
    global accuracy_label, fp_label, fn_label, works_order_label, works_input

    # Set dark theme colors
    background_color = "#330056"  # Dark purple background
    text_color = "#ffffff"  # White text
    
    button_color = "#a86fff"  # Lilac button color
    entry_color = "#ffffff"  # White for entry background
    
    # Use a font that is less likely to appear pixelated and is known for its clean look
    font_family = "Helvetica"
    title_font_size = 24
    button_font_size = 12
    signature_font_size = 8
    
    # Set the initial size and center the window
    window_width = 800
    window_height = 600
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int(screen_width / 2 - window_width / 2)
    center_y = int(screen_height / 2 - window_height / 2)
    root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
    
    # Apply dark theme colors
    root.configure(bg=background_color)
    style = ttk.Style()
    style.theme_use("clam")  # 'clam' theme supports more color customization
    style.configure("TButton", background=button_color, foreground=text_color, font=(font_family, button_font_size), borderwidth=1)
    style.map("TButton", background=[("active", button_color)], foreground=[("active", text_color)])
    
    # Title Label
    title_label = tk.Label(root, text="AI Detection Tool", bg=background_color, fg=text_color, font=(font_family, title_font_size))
    title_label.place(relx=0.5, rely=0.1, anchor="center")
    
    # Entry fields for label thresholds
    label_threshold1 = ttk.Entry(root, background=entry_color, foreground="black")  # Set text color to black
    label_threshold1.insert(0, "Label Threshold 1")
    label_threshold1.bind("<FocusIn>", lambda event: clear_entry(event, label_threshold1, "Label Threshold 1"))
    label_threshold1.bind("<FocusOut>", lambda event: restore_default_text(event, label_threshold1, "Label Threshold 1"))
    label_threshold1.place(relx=0.3, rely=0.2, anchor="center")

    label_threshold2 = ttk.Entry(root, background=entry_color, foreground="black")  # Set text color to black
    label_threshold2.insert(0, "Label Threshold 2")
    label_threshold2.bind("<FocusIn>", lambda event: clear_entry(event, label_threshold2, "Label Threshold 2"))
    label_threshold2.bind("<FocusOut>", lambda event: restore_default_text(event, label_threshold2, "Label Threshold 2"))
    label_threshold2.place(relx=0.3, rely=0.3, anchor="center")

    label_threshold3 = ttk.Entry(root, background=entry_color, foreground="black")  # Set text color to black
    label_threshold3.insert(0, "Label Threshold 3")
    label_threshold3.bind("<FocusIn>", lambda event: clear_entry(event, label_threshold3, "Label Threshold 3"))
    label_threshold3.bind("<FocusOut>", lambda event: restore_default_text(event, label_threshold3, "Label Threshold 3"))
    label_threshold3.place(relx=0.3, rely=0.4, anchor="center")

    # Entry fields for screw thresholds
    screw_threshold1 = ttk.Entry(root, background=entry_color, foreground="black")  # Set text color to black
    screw_threshold1.insert(0, "Screw Threshold 1")
    screw_threshold1.bind("<FocusIn>", lambda event: clear_entry(event, screw_threshold1, "Screw Threshold 1"))
    screw_threshold1.bind("<FocusOut>", lambda event: restore_default_text(event, screw_threshold1, "Screw Threshold 1"))
    screw_threshold1.place(relx=0.5, rely=0.2, anchor="center")

    screw_threshold2 = ttk.Entry(root, background=entry_color, foreground="black")  # Set text color to black
    screw_threshold2.insert(0, "Screw Threshold 2")
    screw_threshold2.bind("<FocusIn>", lambda event: clear_entry(event, screw_threshold2, "Screw Threshold 2"))
    screw_threshold2.bind("<FocusOut>", lambda event: restore_default_text(event, screw_threshold2, "Screw Threshold 2"))
    screw_threshold2.place(relx=0.5, rely=0.3, anchor="center")

    screw_threshold3 = ttk.Entry(root, background=entry_color, foreground="black")  # Set text color to black
    screw_threshold3.insert(0, "Screw Threshold 3")
    screw_threshold3.bind("<FocusIn>", lambda event: clear_entry(event, screw_threshold3, "Screw Threshold 3"))
    screw_threshold3.bind("<FocusOut>", lambda event: restore_default_text(event, screw_threshold3, "Screw Threshold 3"))
    screw_threshold3.place(relx=0.5, rely=0.4, anchor="center")


    label_threshold1_label = tk.Label(root, text="Camera 1 Threshold", bg=background_color, fg=text_color, font=(font_family, 10))
    label_threshold1_label.place(relx=0.1, rely=0.2, anchor="center")

    label_threshold2_label = tk.Label(root, text="Camera 2 Threshold", bg=background_color, fg=text_color, font=(font_family, 10))
    label_threshold2_label.place(relx=0.1, rely=0.3, anchor="center")

    label_threshold3_label = tk.Label(root, text="Camera 3 Threshold", bg=background_color, fg=text_color, font=(font_family, 10))
    label_threshold3_label.place(relx=0.1, rely=0.4, anchor="center")

    works_order_label = tk.Label(root, text="Works Order No.", bg=background_color, fg=text_color, font=(font_family, 10))
    works_order_label.place(relx=0.1, rely=0.475, anchor="center")

    works_input = ttk.Entry(root, background=entry_color, foreground="black")  # Set text color to black
    works_input.place(relx=0.3, rely=0.475, anchor="center")

    # Error Label
    #error_label = tk.Label(root, bg=background_color, fg="red", font=(font_family, 10))
    #error_label.place(relx=0.5, rely=0.7, anchor="center")

    # "Take Photo" Button
    take_photo_button = ttk.Button(root, text="Take Photo", command=on_take_photo_button_click)
    take_photo_button.place(relx=0.5, rely=0.9, anchor="center")

    # Developer Signature at the bottom right
    dev_signature = tk.Label(root, text="Developed by Muhommad Ridhwan Jamsa", bg=background_color, fg=text_color, font=(font_family, signature_font_size))
    dev_signature.place(relx=0.95, rely=0.95, anchor="se")
    
    # Status Indicator
    status_indicator = tk.Label(root, bg="grey", width=50, height=9)  # Make it bigger and more square
    status_indicator.place(relx=0.5, rely=0.75, anchor="center")

    # Accuracy Label
    accuracy_label = ttk.Label(root, text="Accuracy: 100.00%", background=background_color, foreground="yellow",
                            font=(font_family, 14))  # Increase font size
    accuracy_label.place(relx=0.95, rely=0.1, anchor="ne")  # Place in the top right corner

    # False Positive Label
    fp_label = ttk.Label(root, text="FP: 0", background=background_color, foreground="yellow",
                         font=(font_family, 14))  # Increase font size
    fp_label.place(relx=0.95, rely=0.2, anchor="ne")  # Place below the accuracy label

    # False Negative Label
    fn_label = ttk.Label(root, text="FN: 0", background=background_color, foreground="yellow",
                         font=(font_family, 14))  # Increase font size
    fn_label.place(relx=0.95, rely=0.3, anchor="ne")  # Place below the false positive label


    # "Report Mistake" Button
    report_mistake_button = ttk.Button(root, text="Report Mistake", command=report_mistake, style="Orange.TButton")
    report_mistake_button.place(relx=0.5, rely=0.55, anchor="center")  # Adjust placement above the status box

    # "False Positive" Button
    false_positive_button = ttk.Button(root, text="False Positive", command=report_false_positive, style="Orange.TButton")
    false_positive_button.place(relx=0.25, rely=0.55, anchor="center")

    # "False Negative" Button
    false_negative_button = ttk.Button(root, text="False Negative", command=report_false_negative, style="Orange.TButton")
    false_negative_button.place(relx=0.75, rely=0.55, anchor="center")

    # Configure style for the "Report Mistake" button
    style.configure("Orange.TButton", background="#ff8c00", foreground=text_color, font=(font_family, button_font_size),
                borderwidth=1)
    style.map("Orange.TButton", background=[("active", "#ff8c00")], foreground=[("active", text_color)])



    # Create a queue for updating the UI based on the output file
    update_queue = queue.Queue()

    update_ui_based_on_status()

    # Function to process messages from the queue
    def process_queue():
        try:
            while True:  # Process all messages in the queue
                color = update_queue.get_nowait()
                update_callback(color)
        except queue.Empty:
            pass
        root.after(100, process_queue)  # Schedule again in 100 ms

    process_queue()  # Start polling the queue before entering the main loop
    print("Entering the main loop...")
    root.mainloop()
    print("Exited the main loop...")

if __name__ == "__main__":
    main()
